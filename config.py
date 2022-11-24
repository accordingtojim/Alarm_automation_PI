from pickle import TRUE
from tkinter import FALSE
from openpyxl import load_workbook
from openpyxl import Workbook
import os
array_PH=[]     #global variable coming from GUI
array_EH=[]      #global variable coming from GUI
array_HH=[]
array_type=[]     #global variable coming from GUI
array_type_HH=[] 
n_CBESS = []      #global variable coming from GUI
n_other_inverter = []     #global variable coming from GUI
n_battery_rack = []       #global variable coming from GUI
n_battery_bank = []      #global variable coming from GUI
global_list = []
n_HVAC_HH = []
n_HVAC = []
counter = 0
num_HH_GUI = 0
n_PI = 0

def num_assign(i):
    match str(i):
        case "1":
            return 'first'
        case "2":
            return 'second'
        case "3":
            return 'third'
        case "4":
            return 'fourth'
        case "5":
            return 'fifth'
        case "6":
            return 'sixth'
        case "7":
            return 'sixth'
        case "8":
            return 'eight'
        case "9":
            return 'nineth'
        case "10":
            return 'tenth'
def destroy_ww(frame):
    frame.quit()
    frame.destroy()
def array_arrangement(array1,array2):       #example array1= [1,2,3]     #manage if it's not a list but an integer
    c = 0                                   #        array2= [4,5,6]    ->  array3= [4,5,5,6,6,6] then return the sum of elemnts of array3
    array3 = []
    for i in array1:
        for j in range (1,i+1):
            array3.append(array2[c])
        c+=1
    sum=0
    for x in array3:
        sum=sum+x
    return sum
def p_array_arrangement(array1,array2):       #example array1= [1,2,3]
    c = 0                                     #        array2= [4,5,6]    ->  array3= [4,5,5,6,6,6] then return the sum of elemnts of array3
    array3 = []
    for i in array1:
        for j in range (1,i+1):
            array3.append(array2[c])
        c+=1
    return array3
def convert_to_int(array):
    if type(array) == str:
        return int(array)
    else:
        return [int(x) for x in array]
def multiply_array(array1,array2,return_sum):
    array3 = [0] * len(array1) 
    sum = 0
    for i in range(1,len(array1)+1):
        array3[i-1]= array1[i-1]*array2[i-1]
    if return_sum == TRUE:
        for x in array3:
            sum=sum+x
        return sum
    else:
        return array3
def array_name_HVAC(num):
    half = int(num/2)
    array = []
    for i in range(0,half):
        var = '1_'+str(i+1)
        array.append(var)
    for i in range(0,half):
        var ='2_'+str(i+1)
        array.append(var)
    return array
def file_aggregation(lista):
    wb = Workbook()
    ws = wb.active
    for i in lista:
        number_column = 1
        number_row = 1
        wb1 = load_workbook(i)
        ws1 = wb1.worksheets[0]
        while not ((ws1.cell(row=number_row,column=1).value == None) and (ws1.cell(row=number_row + 1,column=1).value == None)) :
            number_row += 1
        number_row -= 1
        rn = ws.max_row
        while not ((ws1.cell(row = 1 ,column = number_column).value == None) and (ws1.cell(row = 1,column = number_column + 1).value == None)) :
            number_column += 1
        number_column -= 1
        cn = ws.max_column
        for i in range (1, number_row + 1):
            for j in range (1, number_column + 1):
                c = ws1.cell(row = i, column = j)
                ws.cell(row = (rn + i), column = (j)).value = c.value
    ws.insert_cols(1)
    ws.insert_cols(1)
    wb.save(filename = 'list_alarm.xlsx')
def file_removal(lista):
    for i in lista:
        os.remove(i)
def file_reorder(path,new_path):
    wb = load_workbook(path)
    ws = wb.active
    number_row = 1
    while not ((ws.cell(row=number_row,column=3).value == None) and (ws.cell(row=number_row + 1,column=3).value  == None) ) :
        number_row += 1
    number_row -= 1
    wb1 = Workbook()
    ws1 = wb1.active
    PI = 'PI0'
    HH = 'HH1HD0'
    counter = 1
    i=1
    j=1
    s=1
    if n_PI != '':
        while str(counter) != str(int(n_PI)+1) :
            for i in range(1,number_row+1):
                if PI + str(counter) in str(ws.cell(row=i,column=3).value):
                    for j in range (1,60):
                        ws1.cell(row=s,column=j).value = ws.cell(row=i,column = 2 + j).value
                    s += 1
            counter += 1
    else:
        while str(counter) != str(int(num_HH_GUI)+1) :
            for i in range(1,number_row+1):
                if (HH + str(counter) in str(ws.cell(row=i,column=3).value)) or (HH + str(counter) in str(ws.cell(row=i,column=6).value)) :
                    for j in range (1,60):
                        ws1.cell(row=s,column=j).value = ws.cell(row=i,column = 2 + j).value
                    s += 1
            counter += 1
    wb1.save (filename = new_path)
def file_header(file_path):
    wb=load_workbook(file_path)
    ws=wb.active
    ws.insert_cols(1)
    ws.insert_cols(1)
    ws.insert_rows(1)
    header = ['ID','Name','Alarm text [en-US], Alarm text 1','FieldInfo [Alarm text 1]','Class','Trigger tag','Trigger bit','Trigger mode','Acknowledgement tag','Acknowledgement bit','Status tag','Status bit','Group','Priority','Single acknowledgement','Info text [en-US], Info text','Additional text 1 [en-US], Alarm text 2','FieldInfo [Alarm text 2]','Additional text 2 [en-US], Alarm text 3','FieldInfo [Alarm text 3]','Additional text 3 [en-US], Alarm text 4','FieldInfo [Alarm text 4]','Additional text 4 [en-US], Alarm text 5','FieldInfo [Alarm text 5]','Additional text 5 [en-US], Alarm text 6','FieldInfo [Alarm text 6]','Additional text 6 [en-US], Alarm text 7','FieldInfo [Alarm text 7]','Additional text 7 [en-US], Alarm text 8','FieldInfo [Alarm text 8]','Additional text 8 [en-US], Alarm text 9','FieldInfo [Alarm text 9]','Additional text 9 [en-US], Alarm text 10','FieldInfo [Alarm text 10]','Alarm parameter 1','Alarm parameter 2','Alarm parameter 3','Alarm parameter 4','Alarm parameter 5','Alarm parameter 6','Alarm parameter 7','Alarm parameter 8','Alarm parameter 9','Alarm parameter 10','Alarm annunciation','Display suppression mask','PLC number','CPU number']
    counter = 1
    for i in header:
        ws.cell(row=1,column=counter).value = i
        counter += 1
    wb.save(filename=file_path)
def file_numbering(file_path):
    wb=load_workbook(file_path)
    ws=wb.active
    counter = 1 
    counter_tot = 0
    counter_ = 0
    while TRUE:
        if ws.cell(row = 2 + counter_tot,column=3).value is None :
                break
        if (("PI0"+str(counter) in ws.cell(row = 2 + counter_tot,column=3).value) or ("Spare" in ws.cell(row = 2 + counter_tot,column=3).value)):
            init_number = 10000*counter 
            ws.cell(row = 2 + counter_tot,column=1).value = init_number + counter_
            ws.cell(row = 2 + counter_tot,column=2).value = init_number + counter_
            counter_tot += 1
            counter_ += 1  
        else: 
            counter += 1
            counter_ = 0       
    wb.save(file_path)
        
    
        


    
            
            
            
        


