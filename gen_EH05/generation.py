from openpyxl import load_workbook
from openpyxl import Workbook
from pickle import FALSE, TRUE
import shutil
import os
import config
def file_creation_1( path_to_template,type_of_template):
    counter = 0
    local_type = config.convert_to_int(config.array_type)
    local_n_EH = config.convert_to_int(config.array_EH)
    local_n_PI = config.convert_to_int(config.n_PI)
    local_n_BB = config.convert_to_int(config.n_battery_bank)
    local_n_BR = config.convert_to_int(config.n_battery_rack)
    local_n_HVAC = config.convert_to_int(config.n_HVAC)
    
    #Creation of new name of the template copy
    match type_of_template:
        case "BB_EH05":         
            path_to_new_template = (path_to_template.replace('.xlsx','')) + '_new_BB' + '.xlsx'
        case "BR_EH05":
            path_to_new_template = (path_to_template.replace('.xlsx','')) + '_new_BR' + '.xlsx'
        case "HVAC_EH05":         
            path_to_new_template = (path_to_template.replace('.xlsx','')) + '_new_HVAC' + '.xlsx'
        case "PEMS_EH05":
            path_to_new_template = (path_to_template.replace('.xlsx','')) + '_new_PEMS' + '.xlsx'    
    
    #Copy the original template to modify the copy and not the original file                
    config.global_list.append(path_to_new_template)
    shutil.copy (path_to_template , path_to_new_template)
    wb = load_workbook(path_to_new_template)    
    ws = wb.active
    
    #Counting rows and columns to, later, make the substitution correctly 
    number_row = 1
    number_column = 1
    while not ((ws.cell(row = 1 ,column = number_column).value == None) and (ws.cell(row = 1,column = number_column + 1).value == None)) :
        number_column += 1
    number_column -= 1
    while not ((ws.cell(row=number_row,column=1).value == None) and (ws.cell(row=number_row + 1,column=1).value == None)) :
        number_row += 1
    number_row -= 1
    
    #In common array
    array_parsed_EH = config.p_array_arrangement(local_type,local_n_EH)
    array_parsed_BB = config.p_array_arrangement(local_type,local_n_BB)
    array_parsed_BR = config.p_array_arrangement(local_type,local_n_BR)
    array_parsed_HVAC = config.p_array_arrangement(local_type,local_n_HVAC)
    array_parsed_Dummy1 = config.p_array_arrangement(local_type,[2])
    array_parsed_Dummy2 = config.p_array_arrangement(local_type,[1])
    #Not in common array
    match type_of_template:
        case "HVAC_EH05":
            array_parsed_tot = config.array_arrangement(array_parsed_EH,array_parsed_HVAC)
            array_parsed_1 = array_parsed_HVAC
            array_parsed_2 = array_parsed_Dummy2
        case "BB_EH05":
            array_parsed_tot=config.array_arrangement(array_parsed_EH,array_parsed_BB)
            array_parsed_1 = array_parsed_BB
            array_parsed_2 = array_parsed_Dummy2
        case "BR_EH05":
            array_parsed_1 = array_parsed_BB
            array_parsed_2 = array_parsed_BR
            array_mid = config.multiply_array(array_parsed_EH,array_parsed_BB,FALSE)
            array_parsed_tot = config.multiply_array(array_mid,array_parsed_BR,TRUE)
        case "PEMS_EH05":
            array_parsed_1 = array_parsed_Dummy1 
            array_parsed_2 = array_parsed_Dummy2
            array_parsed_tot=config.array_arrangement(local_type,local_n_EH)
            
    #In each new file created before, copy and paste all the alarms consequently, in order to modify them rapidly all in one file      
    for n in range(1,array_parsed_tot):
            for j in range(1,number_column+1):
                for i in range(1,number_row+1):
                    ws.cell(row=i+n*number_row,column=j).value=ws.cell(row=i,column=j).value
    #wb.save(path_to_new_template)
    
    for PI in range(1,local_n_PI+1):
        for EH in range(1 , array_parsed_EH[PI-1]+1):
            for VAR1 in range(1 , array_parsed_1[PI-1] + 1):
                number_HVAC = config.array_name_HVAC(array_parsed_1[PI-1]+1)
                for BR in range(1,array_parsed_2[PI-1]+1):
                    for i in range(1,number_row+1):
                        
                        if EH > 9 :
                            name_EH = "EH05HD"
                        else:
                            name_EH = "EH05HD0"
                        if BR > 9 :
                            name_BR = "_BR"
                        else:
                            name_BR = "_BR0"
                        if PI > 9 :
                            name_PI = "PI"
                        else:
                            name_PI = "PI0"
                            
                        match type_of_template:                           
                            case "BR_EH05":
                                string_spare_1 = "|"
                                string_spare_2 = name_BR
                                string_spare_3 = str(BR)
                                string_spare_4 = ","
                                string_spare_5 = "BB0"
                                string_spare_6 = str(VAR1)
                                string_spare_7 = " - "
                                string_spare_8 = "_BMS"
                                string_spare_9 = str(VAR1)
                                string_spare_10 = "_"
                                string_spare_11 = name_BR
                                string_spare_12 =str(BR)
                                string_spare_13 ="."
                            case "BB_EH05":
                                string_spare_1 = "|"
                                string_spare_2 = "BB0"
                                string_spare_3 = str(VAR1)
                                string_spare_4 = "-"
                                string_spare_5 = ""
                                string_spare_6 = ""
                                string_spare_7 = ""
                                string_spare_8 = "_BMS"
                                string_spare_9 = str(VAR1)
                                string_spare_10 = "_"
                                string_spare_11 = ""
                                string_spare_12 = ""
                                string_spare_13 = ""
                            case "PEMS_EH05":
                                array_parsed_tot=config.array_arrangement(local_type,local_n_EH)
                                string_spare_1 = ", PEMS - "
                                string_spare_2 = ""
                                string_spare_3 = ""
                                string_spare_4 = ""
                                string_spare_5 = ""
                                string_spare_6 = ""
                                string_spare_7 = ""
                                string_spare_8 = ""
                                string_spare_9 = ""
                                string_spare_10 = ""
                                string_spare_11 = ""
                                string_spare_12 = ""
                                string_spare_13 = ""
                            case "HVAC_EH05":
                                array_parsed_tot = config.array_arrangement(array_parsed_EH,array_parsed_HVAC)
                                string_spare_1 = "| HVAC"
                                string_spare_2 = number_HVAC[VAR1-1]
                                string_spare_3 = " - "
                                string_spare_4 = ""
                                string_spare_5 = ""
                                string_spare_6 = ""
                                string_spare_7 = ""
                                string_spare_8 = "_HVAC"
                                string_spare_9 = number_HVAC[VAR1-1]
                                string_spare_10 = ""
                                string_spare_11 = ""
                                string_spare_12 = ""
                                string_spare_13 = ""                            
                                
                        if ('Spare' or 'spare') in ws.cell(row=i+counter*number_row,column=1).value:
                            ws.cell(row=i+counter*number_row,column=1).value = str(ws.cell(row =i+counter*number_row, column = 1).value)\
                            + " | "\
                            + str(ws.cell(row =i+counter*number_row, column = 3).value) + "." + str(ws.cell(row =i+counter*number_row, column = 4).value)\
                            + string_spare_1\
                            + string_spare_2\
                            + string_spare_3\
                            + string_spare_4\
                            + string_spare_5\
                            + string_spare_6\
                            + string_spare_7\
                            + name_EH\
                            + str(EH)\
                            + name_PI\
                            + str(PI)   
                            ws.cell(row=i+counter*number_row,column=3).value = name_EH\
                            + str(PI)\
                            + "_"\
                            + str(EH)\
                            + string_spare_8\
                            + string_spare_9\
                            + string_spare_10\
                            + string_spare_11\
                            + string_spare_12\
                            + string_spare_13\
                            + ws.cell(row=i+counter*number_row,column=3).value
                        else:
                            ws.cell(row=i+counter*number_row,column=1).value = str(ws.cell(row =i+counter*number_row, column = 1).value)\
                            + string_spare_1\
                            + string_spare_2\
                            + string_spare_3\
                            + string_spare_4\
                            + string_spare_5\
                            + string_spare_6\
                            + string_spare_7\
                            + name_EH\
                            + str(EH)\
                            + name_PI\
                            + str(PI)
                            ws.cell(row=i+counter*number_row,column=3).value = name_EH\
                            + str(PI)\
                            + "_"\
                            + str(EH)\
                            + string_spare_8\
                            + string_spare_9\
                            + string_spare_10\
                            + string_spare_11\
                            + string_spare_12\
                            + string_spare_13\
                            + ws.cell(row=i+counter*number_row,column=3).value
                    counter += 1
    ws.insert_cols(2)
    wb.save(path_to_new_template)