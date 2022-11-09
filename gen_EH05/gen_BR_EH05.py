from pickle import FALSE, TRUE
from openpyxl import load_workbook
from openpyxl import Workbook
import shutil
import os
import config
def file_creation_6( path_to_template_BR): 
    counter = 0
    local_type = config.convert_to_int(config.array_type)
    local_n_EH = config.convert_to_int(config.array_EH)
    local_n_PI = config.convert_to_int(config.n_PI)
    local_n_BB = config.convert_to_int(config.n_battery_bank)
    local_n_BR = config.convert_to_int(config.n_battery_rack)
    path_to_new_template = (path_to_template_BR.replace('.xlsx','')) + '_new_BR' + '.xlsx'
    config.global_list.append(path_to_new_template)
    shutil.copy (path_to_template_BR , path_to_new_template)
    wb = load_workbook(path_to_new_template)    
    ws = wb.active
    number_row = 1
    number_column = 1
    while not ((ws.cell(row = 1 ,column = number_column).value == None) and (ws.cell(row = 1,column = number_column + 1).value == None)) :
        number_column += 1
    number_column -= 1
    while not ((ws.cell(row=number_row,column=1).value == None) and (ws.cell(row=number_row + 1,column=1).value == None)) :
        number_row += 1
    number_row -= 1
    array_parsed_EH = config.p_array_arrangement(local_type,local_n_EH)
    array_parsed_BB=config.p_array_arrangement(local_type,local_n_BB)
    array_parsed_BR=config.p_array_arrangement(local_type,local_n_BR)
    array_mid = config.multiply_array(array_parsed_EH,array_parsed_BB,FALSE)
    array_parsed_tot = config.multiply_array(array_mid,array_parsed_BR,TRUE)
    for n in range(1,array_parsed_tot):
            for j in range(1,number_column+1):
                for i in range(1,number_row+1):
                    ws.cell(row=i+n*number_row,column=j).value=ws.cell(row=i,column=j).value
    #wb.save(path_to_new_template)
    for PI in range(1 , local_n_PI + 1):
        for EH in range(1 , array_parsed_EH[PI-1]+1):
            for BB in range(1,array_parsed_BB[PI-1]+1):
                for BR in range(1,array_parsed_BR[PI-1]+1):
                    for i in range(1,number_row+1):
                        if ('Spare' or 'spare') in ws.cell(row=i+counter*number_row,column=1).value:
                            ws.cell(row=i+counter*number_row,column=1).value = str(ws.cell(row =i+counter*number_row, column = 1).value)\
                            + " | "\
                            + str(ws.cell(row =i+counter*number_row, column = 3).value) + "." + str(ws.cell(row =i+counter*number_row, column = 4).value)
                            ws.cell(row=i+counter*number_row,column=3).value = "EH05HD0"\
                            + str(EH)\
                            + "_BMS"\
                            + str(BB)\
                            + "_"\
                            + "BR0"\
                            + str(BR)\
                            + "."\
                            + ws.cell(row=i+counter*number_row,column=3).value  
                        else:
                            ws.cell(row=i+counter*number_row,column=1).value = str(ws.cell(row =i+counter*number_row, column = 1).value)\
                            + " | "\
                            + "BR0"\
                            + str(BR)\
                            + ","\
                            + "BB0"\
                            + str(BB)\
                            + " - "\
                            + "EH05HD0"\
                            + str(EH)\
                            + " - PI0"\
                            + str(PI)
                            ws.cell(row=i+counter*number_row,column=3).value = "EH05HD0"\
                            + str(EH)\
                            + "_BMS"\
                            + str(BB)\
                            + "_"\
                            + "BR0"\
                            + str(BR)\
                            + "."\
                            + ws.cell(row=i+counter*number_row,column=3).value
                    counter += 1
    ws.insert_cols(2)
    wb.save(path_to_new_template)