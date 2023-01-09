from openpyxl import load_workbook
from openpyxl import Workbook
import shutil
import os
import config
def file_creation_3( path_to_template_PEMS):
    counter = 0
    local_type = config.convert_to_int(config.array_type)
    local_n_EH = config.convert_to_int(config.array_EH)
    local_n_PI = config.n_PI
    path_to_new_template = (path_to_template_PEMS.replace('.xlsx','')) + '_new_PEMS' + '.xlsx'
    config.global_list.append(path_to_new_template)
    shutil.copy (path_to_template_PEMS , path_to_new_template)
    wb = load_workbook(path_to_new_template)    
    ws = wb.active
    number_row = 1
    number_column = 1
    while not ((ws.cell(row = 1 ,column = number_column).value == None) and (ws.cell(row = 1,column = number_column + 1).value == None)) :
        number_column += 1
    number_column -= 1
    while not ((ws.cell(row=number_row,column=1).value == None) and (ws.cell(row=number_row + 1,column=1).value == None)) :
        number_row += 1
    number_row -= 1
    array_parsed_sum=config.array_arrangement(local_type,local_n_EH)
    array_parsed_EH=config.p_array_arrangement(local_type,local_n_EH)
    for n in range(1,array_parsed_sum):
            for j in range(1,number_column+1):
                for i in range(1,number_row+1):
                    ws.cell(row=i+n*number_row,column=j).value=ws.cell(row=i,column=j).value
    #wb.save(path_to_new_template)
    for PI in range(1 , local_n_PI + 1):
        for EH in range(1,array_parsed_EH[PI-1]+1):
            for PEMS in range(1 , 2):
                for i in range(1,number_row+1):
                    if 'Spare' in ws.cell(row=i+counter*number_row,column=1).value:
                        ws.cell(row=i+counter*number_row,column=1).value = str(ws.cell(row =i+counter*number_row, column = 1).value)\
                        + " | "\
                        + str(ws.cell(row =i+counter*number_row, column = 3).value) + "." + str(ws.cell(row =i+counter*number_row, column = 4).value)\
                        + ", PEMS - EH05HD0"\
                        + str(EH)\
                        +" - PI0"\
                        +str(PI)
                        ws.cell(row=i+counter*number_row,column=3).value = "EH05HD0"\
                        + str(PI)\
                        + "_"\
                        + str(EH)\
                        + str(ws.cell(row=i+counter*number_row,column=3).value)
                    else:
                        ws.cell(row=i+counter*number_row,column=1).value = str(ws.cell(row =i+counter*number_row, column = 1).value)\
                        + ", PEMS - EH05HD0"\
                        + str(EH)\
                        +" - PI0"\
                        +str(PI)
                        ws.cell(row=i+counter*number_row,column=3).value = "EH05HD0"\
                        + str(PI) \
                        + "_"\
                        + str(EH)\
                        + str(ws.cell(row=i+counter*number_row,column=3).value)
                counter += 1
    ws.insert_cols(2)
    wb.save(path_to_new_template)