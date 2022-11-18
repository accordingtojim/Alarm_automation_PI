from openpyxl import load_workbook
from openpyxl import Workbook
import shutil
import os
import config
def file_creation_4( path_to_template_HVAC):
    counter = 0
    local_type = config.convert_to_int(config.array_type)
    local_n_HVAC = config.convert_to_int(config.n_HVAC)
    local_n_EH = config.convert_to_int(config.array_EH)
    local_n_PI = config.convert_to_int(config.n_PI)
    path_to_new_template = (path_to_template_HVAC.replace('.xlsx','')) + '_new_HVAC' + '.xlsx'
    config.global_list.append(path_to_new_template)
    shutil.copy (path_to_template_HVAC , path_to_new_template)
    wb = load_workbook(path_to_new_template)    
    ws = wb.active
    number_row = 1
    number_column = 1
    while not ((ws.cell(row = 1 ,column = number_column).value == None) and (ws.cell(row = 1,column = number_column + 1).value == None)) :
        number_column += 1
    number_column -= 1
    while not ((ws.cell(row=number_row,column=1).value == None) and (ws.cell(row=number_row + 1,column=1).value == None)) :
        number_row += 1
    number_row -= 1
    array_parsed_EH = config.p_array_arrangement(local_type,local_n_EH)
    array_parsed_HVAC = config.p_array_arrangement(local_type,local_n_HVAC)
    array_parsed_sum = config.array_arrangement(array_parsed_EH,array_parsed_HVAC)
    for n in range(1,array_parsed_sum):
            for j in range(1,number_column+1):
                for i in range(1,number_row+1):
                    ws.cell(row=i+n*number_row,column=j).value=ws.cell(row=i,column=j).value
    wb.save(path_to_new_template)
    for PI in range(1 , local_n_PI + 1):      
        for EH in range(1 , array_parsed_EH[PI-1]+1):    #bisogna definire dove attaccare la desinenza del numero di HVAC
            for HVAC in range(1 , array_parsed_HVAC[PI-1]+1):
                number_HVAC = config.array_name_HVAC(array_parsed_HVAC[PI-1]+1)
                for i in range(1, number_row+1):
                    if 'Spare' in ws.cell(row=i+counter*number_row,column=1).value:
                        ws.cell(row=i+counter*number_row,column=1).value = str(ws.cell(row =i+counter*number_row, column = 1).value)\
                        + " | "\
                        + str(ws.cell(row =i+counter*number_row, column = 3).value) + "." + str(ws.cell(row =i+counter*number_row, column = 4).value)
                    else:
                        ws.cell(row=i+counter*number_row,column=1).value = str(ws.cell(row =i+counter*number_row, column = 1).value)\
                        + " | "\
                        + str(ws.cell(row =i+counter*number_row, column = 2).value) + " | HVAC"\
                        + number_HVAC[HVAC-1]\
                        + " - EH05HD0"\
                        + str(EH)\
                        + " - PI0" + str(PI)
                        ws.cell(row=i+counter*number_row,column=4).value = "EH05HD0"\
                        + str(EH) + "_HVAC" + number_HVAC[HVAC-1]\
                        + "_strHMI_" + str(ws.cell(row=i+counter*number_row,column=4).value)  
                counter += 1
    ws.delete_cols(2,1)
    ws.insert_cols(2)
    wb.save(path_to_new_template)