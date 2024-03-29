from openpyxl import load_workbook
from openpyxl import Workbook
import shutil
import os
import config
def file_creation_5( path_to_template_SKID):
    counter = 0
    local_n_PI = config.n_PI
    local_n_PH = config.convert_to_int(config.array_PH)
    local_type = config.convert_to_int(config.array_type)
    path_to_new_template = (path_to_template_SKID.replace('.xlsx','')) + '_new_SKID' + '.xlsx'
    config.global_list.append(path_to_new_template)
    shutil.copy (path_to_template_SKID , path_to_new_template)
    wb = load_workbook(path_to_new_template)    
    ws = wb.active
    number_row = 1
    number_column = 1
    while not ((ws.cell(row = 1 ,column = number_column).value == None) and (ws.cell(row = 1,column = number_column + 1).value == None)) :
        number_column += 1
    number_column -= 1
    while not ((ws.cell(row=number_row,column=1).value == None) and (ws.cell(row=number_row + 1,column=1).value == None)) :
        number_row += 1
    number_row -= 1
    array_parsed_MV = config.p_array_arrangement(local_type,local_n_PH)
    for n in range(1,local_n_PI):
            for j in range(1,number_column+1):
                for i in range(1,number_row+1):
                    ws.cell(row=i+n*number_row,column=j).value=ws.cell(row=i,column=j).value
    for PI in range(1 , local_n_PI + 1):
        for MVskid in range(1,array_parsed_MV[PI-1]+1):
            for i in range(1,number_row+1):
                ws.cell(row=i+counter*number_row,column=1).value = str(ws.cell(row =i+counter*number_row, column = 1).value)\
                + " - "\
                + "PI0"\
                + str(PI)
                ws.cell(row=i+counter*number_row,column=3).value = "PH2HD0"\
                + str(PI)\
                + str(ws.cell(row=i+counter*number_row,column=3).value)
            counter += 1
    ws.insert_cols(2)
    wb.save(path_to_new_template)