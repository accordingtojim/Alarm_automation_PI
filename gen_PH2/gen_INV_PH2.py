from openpyxl import load_workbook
from openpyxl import Workbook
import shutil
import os
import config
number_of_column = 2
def file_creation_0( path_to_template_CL):
    counter = 0
    local_type = config.convert_to_int(config.array_type)
    local_n_PH = config.convert_to_int(config.array_PH)
    local_n_PI = config.n_PI
    local_n_CBESS = config.convert_to_int(config.n_CBESS)
    path_to_new_template = (path_to_template_CL.replace('template_CL.xlsx','')) + '_new_CL' + '.xlsx'
    config.global_list.append(path_to_new_template)
    shutil.copy (path_to_template_CL , path_to_new_template)
    wb = load_workbook(path_to_new_template)    
    ws = wb.active
    number_row = 1
    number_column = 1
    #conto il numero di righe e colonne con ciclo while, la funziona ws.max_column a volte non mi dava lo stesso risultato
    while not ((ws.cell(row = 1 ,column = number_column).value == None) and (ws.cell(row = 1,column = number_column + 1).value == None)) :
        number_column += 1
    number_column -= 1
    while not ((ws.cell(row=number_row,column=1).value == None) and (ws.cell(row=number_row + 1,column=1).value == None)) :
        number_row += 1
    number_row -= 1
    #genero in un excel dedicato il numero di copie del file di template che mi servono
    array_parsed_PH = config.p_array_arrangement(local_type,local_n_PH)
    array_parsed_CBESS=config.p_array_arrangement(local_type,local_n_CBESS)
    array_parsed_tot=config.array_arrangement(array_parsed_PH,array_parsed_CBESS)
    for n in range(1,array_parsed_tot * number_of_column):     #number_of_column è inizializzato a 2
            for j in range(1,number_column+1):
                for i in range(1,number_row+1):
                    ws.cell(row=i+n*number_row,column=j).value=ws.cell(row=i,column=j).value
    #wb.save(path_to_new_template)
    #nel file excel dedicato vado a sostituire e riordinare le colonne per avere la stessa struttura che ha un excel esportato da TIA
    for PI in range(1,local_n_PI+1):    
        for PH in range(1 ,array_parsed_PH[PI-1]+1):
            for C_BESS in range(1,array_parsed_CBESS[PI-1]+1):
                for COLUMN in range(1 , number_of_column + 1):
                    for i in range(1,number_row+1):
                        if 'Spare' in ws.cell(row=i+counter*number_row,column=1).value:
                            ws.cell(row=i+counter*number_row,column=1).value = str(ws.cell(row =i+counter*number_row, column = 1).value)\
                            + " | "\
                            + str(ws.cell(row =i+counter*number_row, column = 3).value) + "." + str(ws.cell(row =i+counter*number_row, column = 4).value)\
                            + " | CL"\
                            + str(COLUMN)\
                            + ",PCS"\
                            + str(C_BESS)\
                            + " - "\
                            + "PH2HD0"\
                            + str(PH)\
                            + " - PI0"\
                            +str(PI)
                            ws.cell(row=i+counter*number_row,column=3).value = "PH2HD0"\
                            + str(PI)\
                            + "_PCS"\
                            + str(C_BESS)\
                            + "_CL"\
                            + str(COLUMN)\
                            + "_Status_HMI."\
                            + str(ws.cell(row=i+counter*number_row,column=3).value)
                        else:
                            ws.cell(row=i+counter*number_row,column=1).value = str(ws.cell(row =i+counter*number_row, column = 1).value)\
                            + " | CL"\
                            + str(COLUMN)\
                            + ",PCS"\
                            + str(C_BESS)\
                            + " - "\
                            + "PH2HD0"\
                            + str(PH)\
                            + " - PI0"\
                            +str(PI)
                            ws.cell(row=i+counter*number_row,column=3).value = "PH2HD0"\
                            + str(PI)\
                            + "_"\
                            + "PCS"\
                            + str(C_BESS)\
                            + "_CL"\
                            + str(COLUMN)\
                            + "_Status_HMI."\
                            + str(ws.cell(row=i+counter*number_row,column=3).value)
                    counter += 1
    ws.insert_cols(2)
    wb.save(path_to_new_template)